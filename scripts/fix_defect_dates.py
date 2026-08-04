"""Fix Defect Tracker date columns (Submitted + Action Taken) to d-mmm-yyyy."""
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils.datetime import from_excel

DOCS = Path(__file__).resolve().parents[1] / "docs" / "Milestone_Documents"
DATE_FMT = "d-mmm-yyyy"


def fix_defect_dates(path: Path) -> Path:
    wb = openpyxl.load_workbook(path)
    ws = wb["Defects"]
    ref = ws.cell(2, 3)

    for row in range(2, ws.max_row + 1):
        for col in (3, 9):  # Submitted Date, Action Taken Date
            cell = ws.cell(row, col)
            v = cell.value
            if v is None:
                continue
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                try:
                    cell.value = from_excel(v)
                except Exception:
                    continue
            if isinstance(cell.value, datetime):
                cell.number_format = ref.number_format or DATE_FMT

    try:
        wb.save(path)
        out = path
    except PermissionError:
        out = path.with_name(path.stem + "_FIXED.xlsx")
        wb.save(out)
        print(f"LOCKED: saved {out.name} — close Excel and rename to {path.name}")
    else:
        print(f"Fixed dates -> {path.name}")
    return out


if __name__ == "__main__":
    fix_defect_dates(DOCS / "Defect_Tracker_v1_Filled.xlsx")
