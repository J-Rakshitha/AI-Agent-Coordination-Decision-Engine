"""Apply M1/M2 Excel formatting to Milestone 3 rows (row heights, borders, wrap)."""
from copy import copy
import openpyxl

ROOT = r"c:\Users\akula\OneDrive\Desktop\new infosys\ai-agent-coordination-engine-phase1.final\ai-agent-cursor_project_phase1.final\ai-agent-coordination-engine"
DOCS = f"{ROOT}/docs/Milestone_Documents"
FILES = [
    f"{DOCS}/Agile_Template_v1_Filled.xlsx",
    f"{DOCS}/Defect_Tracker_v1_Filled.xlsx",
    f"{DOCS}/Unit_Test_Plan_v1_Filled.xlsx",
]


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def copy_row_style(ws, src_row, dst_row, max_col):
    src_h = ws.row_dimensions[src_row].height
    if src_h:
        ws.row_dimensions[dst_row].height = src_h
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def fix_agile(path):
    wb = openpyxl.load_workbook(path)

    # Product Backlog: rows 22-33 — match Sprint 1/2 (54 or 69.75 for long stories)
    pb = wb["Product Backlog"]
    long_pb_rows = {22, 25, 30, 31, 32, 33, 34, 35, 36, 37}
    for r in range(22, 40):
        ref = 21 if r in long_pb_rows else 2
        copy_row_style(pb, ref, r, 8)

    # Sprint Backlog: header row 31 from row 20; tasks 32-48 from row 21
    sb = wb["Sprint Backlog"]
    copy_row_style(sb, 20, 31, 23)
    sb.cell(31, 1).value = "SPRINT  3  BACKLOG"
    if "A31:W31" not in [str(m) for m in sb.merged_cells.ranges]:
        sb.merge_cells("A31:W31")
    long_task_rows = {33, 36, 39, 41, 43, 44, 45, 46, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58}
    for r in range(32, 59):
        ref = 22 if r in long_task_rows else 21
        copy_row_style(sb, ref, r, 23)

    # Stand up: rows 9-14
    su = wb["Stand up Meeting"]
    long_su = {9, 11, 12, 13, 14}
    for r in range(9, 15):
        ref = 3 if r in long_su else 7
        copy_row_style(su, ref, r, 7)

    # Retrospection: row 4 from row 3
    retro = wb["Retrospection"]
    copy_row_style(retro, 3, 4, 9)

    wb.save(path)
    print(f"Fixed Agile: {path}")


def fix_defects(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Defects"]
    long_rows = {7, 8, 9, 10, 13}
    for r in range(8, 14):
        ref = 7 if r in long_rows else 2
        copy_row_style(ws, ref, r, 11)
    wb.save(path)
    print(f"Fixed Defects: {path}")


def fix_unit_tests(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["UT"]
    long_rows = {23, 24, 25, 27, 28, 29, 30, 31, 35, 36, 37, 40, 41, 42, 43, 44, 45, 46, 47}
    last = ws.max_row
    for r in range(23, last + 1):
        if ws.cell(r, 1).value is None:
            continue
        ref = 22 if r in long_rows else 2
        copy_row_style(ws, ref, r, 6)
    wb.save(path)
    print(f"Fixed Unit Tests: {path}")


if __name__ == "__main__":
    for f in FILES:
        if "Agile" in f:
            fix_agile(f)
        elif "Defect" in f:
            fix_defects(f)
        elif "Unit_Test" in f:
            fix_unit_tests(f)
    print("Done — formatting applied to all files.")
