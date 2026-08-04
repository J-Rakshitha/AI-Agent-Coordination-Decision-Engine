"""Fix Sprint 3 dates, defect order/format, consolidate to final documents only."""
from copy import copy
from datetime import datetime
import zipfile
import openpyxl
from pathlib import Path

DOCS = Path(__file__).resolve().parents[1] / "docs" / "Milestone_Documents"
DATE_FMT = "d-mmm-yyyy"

RETRO_S3_START = datetime(2026, 7, 27)
RETRO_S3_END = datetime(2026, 8, 4)

TASK_DATES = [
    ("T025", datetime(2026, 7, 27), datetime(2026, 7, 28)),
    ("T026", datetime(2026, 7, 28), datetime(2026, 7, 29)),
    ("T027", datetime(2026, 7, 29), datetime(2026, 7, 30)),
    ("T028", datetime(2026, 7, 30), datetime(2026, 7, 31)),
    ("T029", datetime(2026, 7, 31), datetime(2026, 8, 1)),
    ("T030", datetime(2026, 8, 1), datetime(2026, 8, 2)),
    ("T031", datetime(2026, 8, 2), datetime(2026, 8, 2)),
    ("T032", datetime(2026, 8, 2), datetime(2026, 8, 3)),
    ("T033", datetime(2026, 8, 3), datetime(2026, 8, 3)),
    ("T034", datetime(2026, 8, 3), datetime(2026, 8, 4)),
    ("T035", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T036", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T037", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T038", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T039", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T040", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T041", datetime(2026, 8, 4), datetime(2026, 8, 4)),
]

# Sprint 3 defects in chronological order (by submitted date)
DEFECTS_S3 = [
    {
        "sl": 8,
        "submitted": datetime(2026, 7, 28),
        "description": "GitHub webhook POST returned HTTP 401 Invalid signature — ngrok reachable but HMAC verification failed.",
        "action": "Aligned GitHub repo webhook secret with GITHUB_WEBHOOK_SECRET in backend/.env via API PATCH; added signature to pytest.",
        "type": "Logical",
        "remarks": "Recent Deliveries shows 200 OK; manual and webhook sync both work.",
    },
    {
        "sl": 10,
        "submitted": datetime(2026, 7, 30),
        "description": "google-genai SDK returned 429 quota exceeded for gemini-2.0-flash with AQ auth key.",
        "action": "Switched default LLM_MODEL to gemini-flash-latest in config.py and .env; verified used_llm=True in live test.",
        "type": "Others",
        "remarks": "LLM badge appears on Code Review and Tool Selector agents in demo.",
    },
    {
        "sl": 7,
        "submitted": datetime(2026, 8, 1),
        "description": "GET /api/incidents/ returned HTTP 500 on existing SQLite database — OperationalError: no such column incidents.sla_minutes.",
        "action": "Created Alembic migration 001_add_incident_sla_fields and documented 'alembic upgrade head' in README.",
        "type": "Logical",
        "remarks": "Incidents API returns 200; SLA fields visible on AIOps cards.",
    },
    {
        "sl": 9,
        "submitted": datetime(2026, 8, 3),
        "description": "Dev-Collab 'Simulate Conflict' displayed misleading 'Backend not reachable' error after ~10 seconds.",
        "action": "Increased simulateDemoConflict and simulateIncident axios timeout from 10s to 45s; improved error message for ECONNABORTED.",
        "type": "User Interface",
        "remarks": "Simulate Conflict completes in ~15s and shows conflict + notifications.",
    },
    {
        "sl": 11,
        "submitted": datetime(2026, 8, 4),
        "description": "Discord webhook test returned 401 when webhook URL was revoked and regenerated in channel settings.",
        "action": "Updated DISCORD_WEBHOOK_URL in .env and re-ran POST /api/system/test-discord-webhook.",
        "type": "Others",
        "remarks": "Discord #ai-alerts receives Dev-Collab and AIOps alerts.",
    },
]


def pick_source(name: str) -> Path:
    dated = DOCS / name.replace(".xlsx", "_dated.xlsx")
    standard = DOCS / name
    if dated.exists() and standard.exists():
        return dated if dated.stat().st_mtime >= standard.stat().st_mtime else standard
    return dated if dated.exists() else standard


def set_date_cell(cell, dt: datetime, ref_cell):
    cell.value = dt
    cell.number_format = ref_cell.number_format or DATE_FMT


def fix_agile(src: Path, dst: Path):
    wb = openpyxl.load_workbook(src)
    retro = wb["Retrospection"]
    ref = retro.cell(2, 3)
    set_date_cell(retro.cell(4, 3), RETRO_S3_START, ref)
    set_date_cell(retro.cell(4, 4), RETRO_S3_END, ref)

    sb = wb["Sprint Backlog"]
    task_map = {tid: (s, e) for tid, s, e in TASK_DATES}
    date_ref = sb.cell(21, 4)
    for row in range(32, 49):
        tid = sb.cell(row, 2).value
        if tid in task_map:
            start, end = task_map[tid]
            set_date_cell(sb.cell(row, 4), start, date_ref)
            set_date_cell(sb.cell(row, 5), end, date_ref)

    wb.save(dst)
    print("Fixed Agile ->", dst.name)


def fix_defects(src: Path, dst: Path):
    wb = openpyxl.load_workbook(src)
    ws = wb["Defects"]
    ref_sub = ws.cell(2, 3)
    ref_act = ws.cell(2, 9)

    # Normalize date format on all existing rows
    for row in range(2, ws.max_row + 1):
        for col in (3, 9):
            cell = ws.cell(row, col)
            if isinstance(cell.value, datetime):
                cell.number_format = DATE_FMT
            elif isinstance(cell.value, (int, float)) and col == 9:
                # Excel serial numbers from bad import
                from openpyxl.utils.datetime import from_excel
                try:
                    cell.value = from_excel(cell.value)
                    cell.number_format = DATE_FMT
                except Exception:
                    cell.value = None

    # Rewrite Sprint 3 rows in chronological order (rows 8-12)
    start_row = 8
    for i, d in enumerate(DEFECTS_S3):
        row = start_row + i
        ws.cell(row, 1, d["sl"])
        ws.cell(row, 2, "J Rakshitha")
        set_date_cell(ws.cell(row, 3), d["submitted"], ref_sub)
        ws.cell(row, 4, d["description"])
        ws.cell(row, 5, "Sprint 3")
        ws.cell(row, 6, "J Rakshitha")
        ws.cell(row, 7, d["type"])
        ws.cell(row, 8, d["action"])
        set_date_cell(ws.cell(row, 9), d["submitted"], ref_act)  # action same day as submit
        ws.cell(row, 10, "Closed")
        ws.cell(row, 11, d["remarks"])

    wb.save(dst)
    print("Fixed Defects ->", dst.name)


def consolidate():
    finals = [
        "Agile_Template_v1_Filled.xlsx",
        "Defect_Tracker_v1_Filled.xlsx",
        "Unit_Test_Plan_v1_Filled.xlsx",
    ]
    tmp_ag = DOCS / "_tmp_agile.xlsx"
    tmp_df = DOCS / "_tmp_defects.xlsx"

    fix_agile(pick_source(finals[0]), tmp_ag)
    fix_defects(pick_source(finals[1]), tmp_df)

    # Unit test: copy as-is from existing
    ut_src = pick_source(finals[2])
    ut_dst = DOCS / finals[2]

    for tmp, final_name in [(tmp_ag, finals[0]), (tmp_df, finals[1])]:
        dst = DOCS / final_name
        try:
            if dst.exists():
                dst.unlink()
            tmp.rename(dst)
        except PermissionError:
            alt = DOCS / final_name.replace(".xlsx", " (Latest).xlsx")
            tmp.rename(alt)
            print("LOCKED:", final_name, "-> use", alt.name)

    # Remove duplicates
    for p in DOCS.glob("*"):
        if p.name.startswith("~") or p.name.endswith(".zip"):
            continue
        if p.name in finals:
            continue
        if p.name.endswith("_dated.xlsx") or p.name.endswith(" (Latest).xlsx") or p.name.startswith("_tmp"):
            try:
                p.unlink()
                print("Removed extra:", p.name)
            except Exception:
                pass

    # Rebuild zip
    zip_path = DOCS / "Milestone_Documents_M1_M2_M3.zip"
    zip_path.unlink(missing_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in finals:
            p = DOCS / name
            latest = DOCS / name.replace(".xlsx", " (Latest).xlsx")
            src = latest if latest.exists() and not p.exists() else p
            if latest.exists() and p.exists():
                src = latest if latest.stat().st_mtime > p.stat().st_mtime else p
            zf.write(src, name)
    print("Rebuilt zip")


if __name__ == "__main__":
    consolidate()
