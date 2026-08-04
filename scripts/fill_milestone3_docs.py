"""Fill Agile, Defect Tracker, and Unit Test Plan with Milestone 3 (Sprint 3) data."""
from datetime import datetime
from copy import copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

ROOT = r"c:\Users\akula\OneDrive\Desktop\new infosys\ai-agent-coordination-engine-phase1.final\ai-agent-cursor_project_phase1.final\ai-agent-coordination-engine"
ASSIGNEE = "J Rakshitha"

# --- Milestone 3 Product Backlog (Sprint 3) ---
PRODUCT_BACKLOG_M3 = [
    ("Sprint 3", "Sprint 3", "US021", "As a system, I want a Notification Agent delivering multi-channel team alerts (WebSocket, Gmail SMTP, Slack, Discord) so that conflicts and incidents reach the team instantly.", "Must Have", "US020", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US022", "As a user, I want a Team Notifications panel on the Overview page so that I can audit all alert deliveries across channels.", "Must Have", "US021", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US023", "As a user, I want Code Review Agent output visible on conflict cards so that merge-risk advice is actionable in the UI.", "Must Have", "US007", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US024", "As a system, I want full agent decision logging across Dev-Collab and AIOps pipelines so that every step is explainable.", "Must Have", "US011", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US025", "As a QA engineer, I want Milestone 3 coordination and memory tests so that cross-module linking and knowledge base behaviour are verified.", "Must Have", "US024", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US026", "As a system, I want a GitHub webhook endpoint for real-time PR events so that conflicts sync without manual button clicks.", "Must Have", "US013", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US027", "As a system, I want Alembic DB migrations with SLA fields on incidents so that schema updates do not require deleting the database.", "Must Have", "US008", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US028", "As a user, I want a live SLA countdown timer on AIOps incident cards so that SLA compliance is visible during demos.", "Should Have", "US027", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US029", "As a developer, I want the Hybrid AI Client migrated to the google-genai SDK so that AQ and AIza Gemini keys both work.", "Must Have", "US004", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US030", "As a QA engineer, I want enterprise integration tests (webhook, Slack, Discord, Gmail, SLA) so that external channels are regression-safe.", "Must Have", "US021, US026", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US031", "As a developer, I want a shared GitHub sync service used by manual sync and webhooks so that conflict detection logic is not duplicated.", "Must Have", "US026", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US032", "As a developer, I want README and .env.example updated for Milestone 3 enterprise features so that the project is submission-ready.", "Should Have", "US030", ASSIGNEE, "Completed"),
]

# --- Sprint 3 tasks (T025+) ---
SPRINT3_TASKS = [
    ("US021", "T025", "Implement Notification Agent multi-channel delivery (Slack incoming webhook)", datetime(2026, 7, 24), datetime(2026, 7, 25), "Development", "Completed", 4, 4, 0),
    ("US021", "T026", "Add Discord webhook channel + Gmail SMTP configurable team recipients", datetime(2026, 7, 25), datetime(2026, 7, 26), "Development", "Completed", 5, 5, 0),
    ("US022", "T027", "Build Team Notifications panel component on Overview page", datetime(2026, 7, 26), datetime(2026, 7, 27), "Development", "Completed", 4, 4, 0),
    ("US023", "T028", "Display Code Review Agent notes on Dev-Collab conflict cards", datetime(2026, 7, 27), datetime(2026, 7, 28), "Development", "Completed", 4, 0),
    ("US024", "T029", "Expand incident pipeline decision logging for all agents (Severity, Tool, Notification)", datetime(2026, 7, 28), datetime(2026, 7, 29), "Development", "Completed", 4, 4, 0),
    ("US026", "T030", "Implement GitHub webhook route with HMAC signature verification", datetime(2026, 7, 29), datetime(2026, 7, 30), "Development", "Completed", 5, 5, 0),
    ("US031", "T031", "Extract shared github_sync_service for manual sync and webhook triggers", datetime(2026, 7, 30), datetime(2026, 7, 30), "Integration", "Completed", 3, 0),
    ("US027", "T032", "Set up Alembic migrations and add SLA fields (sla_minutes, sla_deadline, escalated_to)", datetime(2026, 7, 31), datetime(2026, 8, 1), "Development", "Completed", 5, 5, 0),
    ("US028", "T033", "Build SlaCountdown live timer component on AIOps incident cards", datetime(2026, 8, 1), datetime(2026, 8, 2), "Development", "Completed", 4, 4, 0),
    ("US029", "T034", "Migrate Hybrid AI Client from LangChain to google-genai SDK (AQ + AIza keys)", datetime(2026, 8, 2), datetime(2026, 8, 3), "Development", "Completed", 5, 5, 0),
    ("US030", "T035", "Add GET /api/system/integrations and test endpoints (email, Discord, Teams)", datetime(2026, 8, 3), datetime(2026, 8, 3), "Development", "Completed", 3, 0),
    ("US025", "T036", "Write Milestone 3 tests: cross-module linking, memory, notifications, code review", datetime(2026, 8, 3), datetime(2026, 8, 4), "Testing", "Completed", 5, 5, 0),
    ("US030", "T037", "Write enterprise integration tests (webhook, Slack, Discord, Gmail, SLA) — 14 tests", datetime(2026, 8, 4), datetime(2026, 8, 4), "Testing", "Completed", 6, 6, 0),
    ("US029", "T038", "Write LLM hybrid tests for google-genai SDK fallback behaviour — 3 tests", datetime(2026, 8, 4), datetime(2026, 8, 4), "Testing", "Completed", 3, 0),
    ("US032", "T039", "Fix Simulate Conflict/Incident frontend timeout (45s) and clearer error messages", datetime(2026, 8, 4), datetime(2026, 8, 4), "Development", "Completed", 2, 0),
    ("US026", "T040", "Configure ngrok + GitHub webhook secret sync; verify live PR delivery", datetime(2026, 8, 4), datetime(2026, 8, 4), "Integration", "Completed", 3, 0),
    ("US032", "T041", "Update README, .env.example with M3 features; full regression (46+ tests)", datetime(2026, 8, 4), datetime(2026, 8, 4), "Documentation", "Completed", 3, 0),
]

STANDUP_M3 = [
    ("Sprint 3", "Day 3", "POST /api/incidents/ returned HTTP 500 — SQLite DB missing new SLA columns after model update.", "Ran 'alembic upgrade head' to apply migration 001_add_incident_sla_fields without deleting existing data."),
    ("Sprint 3", "Day 5", "GitHub webhook deliveries returned HTTP 401 — signature mismatch between GitHub secret and backend .env.", "Synced GITHUB_WEBHOOK_SECRET via GitHub API PATCH and verified signed ping returns 200 through ngrok."),
    ("Sprint 3", "Day 7", "Simulate Conflict button showed 'Backend not reachable' although FastAPI was running on port 8000.", "Root cause was axios 10s timeout; LLM + Slack/Discord/Gmail pipeline takes 15-20s. Increased timeout to 45s in apiClient.js."),
    ("Sprint 3", "Day 9", "Microsoft Teams webhook setup failed — personal Gmail Microsoft account rejected by Power Automate (AADSTS500200).", "Documented Teams as optional (M365 only); completed Slack + Discord + Gmail as live multi-channel demo instead."),
]

RETRO_M3 = (
    3, "Sprint 3", datetime(2026, 7, 27), datetime(2026, 8, 4), ASSIGNEE,
    "Testing external integrations (Slack, Discord, Gmail, GitHub webhook) early with dedicated /api/system/test-* endpoints before UI demo.",
    "Assuming 10s frontend timeout is enough when agents run LLM + three notification channels sequentially.",
    "Multi-channel Notification Agent pattern — one simulate action alerts Slack, Discord, Gmail and dashboard in parallel.",
    "Delivered Milestone 3 with 46+ automated tests, real GitHub webhook, and enterprise polish (Alembic, SLA UI, google-genai SDK).",
)

DEFECTS_M3 = [
    (7, "J Rakshitha", datetime(2026, 8, 1), "GET /api/incidents/ returned HTTP 500 on existing SQLite database — OperationalError: no such column incidents.sla_minutes.", "Sprint 3", "J Rakshitha", "Logical", "Created Alembic migration 001_add_incident_sla_fields and documented 'alembic upgrade head' in README.", datetime(2026, 8, 1), "Closed", "Incidents API returns 200; SLA fields visible on AIOps cards."),
    (8, "J Rakshitha", datetime(2026, 7, 28), "GitHub webhook POST returned HTTP 401 Invalid signature — ngrok reachable but HMAC verification failed.", "Sprint 3", "J Rakshitha", "Logical", "Aligned GitHub repo webhook secret with GITHUB_WEBHOOK_SECRET in backend/.env via API PATCH; added signature to pytest.", datetime(2026, 7, 28), "Closed", "Recent Deliveries shows 200 OK; manual and webhook sync both work."),
    (9, "J Rakshitha", datetime(2026, 8, 3), "Dev-Collab 'Simulate Conflict' displayed misleading 'Backend not reachable' error after ~10 seconds.", "Sprint 3", "J Rakshitha", "User Interface", "Increased simulateDemoConflict and simulateIncident axios timeout from 10s to 45s; improved error message for ECONNABORTED.", datetime(2026, 8, 3), "Closed", "Simulate Conflict completes in ~15s and shows conflict + notifications."),
    (10, "J Rakshitha", datetime(2026, 7, 30), "google-genai SDK returned 429 quota exceeded for gemini-2.0-flash with AQ auth key.", "Sprint 3", "J Rakshitha", "Others", "Switched default LLM_MODEL to gemini-flash-latest in config.py and .env; verified used_llm=True in live test.", datetime(2026, 7, 30), "Closed", "LLM badge appears on Code Review and Tool Selector agents in demo."),
    (11, "J Rakshitha", datetime(2026, 8, 4), "Discord webhook test returned 401 when webhook URL was revoked and regenerated in channel settings.", "Sprint 3", "J Rakshitha", "Others", "Updated DISCORD_WEBHOOK_URL in .env and re-ran POST /api/system/test-discord-webhook.", datetime(2026, 8, 4), "Closed", "Discord #ai-alerts receives Dev-Collab and AIOps alerts."),
]

UNIT_TESTS_M3 = [
    (22, "Cross-module coordination links incident to commit", "Start two edit sessions on checkout.py, check-conflicts, suggest-resolution, then POST /api/incidents/ingest-metrics for checkout-service with severe metrics.", "A resolved Dev-Collab conflict exists before the production incident.", "Incident response includes linked_commit with file_path=checkout.py and had_conflict=true.", "Pass — cross-module link confirmed in pytest."),
    (23, "Decision log records both Dev-Collab and AIOps modules", "Simulate conflict, suggest resolution, simulate incident, then GET /api/system/decision-log.", "Both pipelines have run at least once.", "Decision log contains dev_collab and aiops modules and agents including Notification, Tool Selector, Root-Cause.", "Pass — all expected agents logged."),
    (24, "Short-term memory builds across repeated incidents", "POST same severe ingest-metrics payload twice for auth-service; check decision-log and knowledge-base.", "Identical anomaly signature submitted twice.", "Knowledge base has one auth-service entry with success_count=2.", "Pass — long-term memory reinforced."),
    (25, "Conflict resolution increments knowledge base success_count", "Run two full conflict cycles on payment_service.py:validateLogin with session end between cycles.", "Same file/function collision occurs twice.", "Knowledge entry key_signature=payment_service.py:validateLogin with success_count=2.", "Pass — conflict pattern learned."),
    (26, "Code Review Agent runs on conflict detection", "Start overlapping edit sessions on checkout.py and POST check-conflicts; GET /api/dev-collab/conflicts.", "Overlap detected on same file/function.", "Conflict record includes non-empty code_review_notes field.", "Pass — Code Review pipeline verified."),
    (27, "Notification Agent persists team alerts", "Simulate conflict, resolve it, simulate incident; GET /api/system/notifications.", "Notification Agent enabled.", "Notifications include conflict_detected, conflict_resolved, incident_created with websocket and email channels.", "Pass — audit trail persisted."),
    (28, "Incident includes SLA fields after migration", "POST /api/incidents/simulate; GET /api/incidents/.", "Alembic migration applied.", "Incident has sla_minutes and sla_deadline populated.", "Pass — SLA fields present."),
    (29, "GitHub status includes webhook configuration info", "GET /api/dev-collab/github/status.", "GitHub integration configured in .env.", "Response includes webhook_url ending in /api/dev-collab/github/webhook and webhook_secret_configured flag.", "Pass — webhook metadata exposed."),
    (30, "GitHub webhook ping with valid HMAC signature", "POST /api/dev-collab/github/webhook with ping payload and valid X-Hub-Signature-256 header.", "GITHUB_WEBHOOK_SECRET set in test.", "HTTP 200, processed=true, action=ping.", "Pass — signature verification accepts valid requests."),
    (31, "GitHub webhook rejects bad signature in production", "POST webhook with invalid signature when ENV=production.", "Production mode with secret configured.", "HTTP 401 Unauthorized.", "Pass — invalid signatures rejected."),
    (32, "Slack notification recorded when webhook configured", "Monkeypatch SLACK_WEBHOOK_URL; simulate-demo-conflict; GET notifications.", "Slack webhook URL configured.", "Notifications channel set includes 'slack'.", "Pass — Slack delivery logged."),
    (33, "Discord notification recorded when webhook configured", "Monkeypatch DISCORD_WEBHOOK_URL; simulate-demo-conflict; GET notifications.", "Discord webhook URL configured.", "Notifications channel set includes 'discord'.", "Pass — Discord delivery logged."),
    (34, "Gmail email notification when SMTP configured", "Monkeypatch SMTP settings and team emails; simulate-demo-conflict; GET notifications.", "Gmail SMTP credentials configured.", "Notifications include channel 'email' to configured recipient.", "Pass — real email path exercised in test."),
    (35, "Hybrid AI client uses LLM when GenAI succeeds", "Mock google.genai.Client to return text; call HybridAIClient.reason with fallback.", "LLM_ENABLED=true and valid API key.", "used_llm=true and LLM text returned.", "Pass — google-genai SDK path verified."),
    (36, "Hybrid AI client falls back on simulated failure", "Set FORCE_SIMULATED_FAILURE=true; call HybridAIClient.reason.", "LLM failure simulation enabled.", "used_llm=false and rule-based fallback text returned.", "Pass — demo never crashes on LLM outage."),
    (37, "Hybrid AI client falls back when GenAI raises", "Mock generate_content to raise RuntimeError; call HybridAIClient.reason.", "GenAI client throws authentication error.", "used_llm=false, fallback text returned, error captured.", "Pass — resilient fallback confirmed."),
    (38, "Integrations status endpoint lists all channels", "GET /api/system/integrations.", "Backend running.", "Response includes email, slack, discord, teams, github, llm sections.", "Pass — integration dashboard API works."),
    (39, "Test email endpoint requires SMTP configuration", "POST /api/system/test-email with empty SMTP password.", "SMTP not fully configured.", "HTTP 400 with configuration guidance.", "Pass — guardrail prevents silent failure."),
    (40, "Full regression suite passes after Milestone 3", "Run 'python -m pytest -v' from backend directory.", "All M1+M2+M3 tests present.", "46+ tests pass with no failures.", "Pass — full suite green."),
]


def write_row(ws, row, values, start_col=1):
    for i, val in enumerate(values, start=start_col):
        ws.cell(row=row, column=i, value=val)


def copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def copy_row_style(ws, src_row, dst_row, max_col):
    src_h = ws.row_dimensions[src_row].height
    if src_h:
        ws.row_dimensions[dst_row].height = src_h
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def apply_m3_formatting(wb):
    """Match Milestone 1/2 row heights, borders, and text-wrap on Milestone 3 rows."""
    pb = wb["Product Backlog"]
    long_pb = {22, 25, 30, 31, 32, 33}
    for r in range(22, 34):
        copy_row_style(pb, 21 if r in long_pb else 2, r, 8)

    sb = wb["Sprint Backlog"]
    copy_row_style(sb, 20, 31, 23)
    sb.cell(31, 1).value = "SPRINT  3  BACKLOG"
    if "A31:W31" not in [str(m) for m in sb.merged_cells.ranges]:
        sb.merge_cells("A31:W31")
    long_task = {33, 36, 39, 41, 43, 44, 45, 46}
    for r in range(32, 49):
        copy_row_style(sb, 22 if r in long_task else 21, r, 23)

    su = wb["Stand up Meeting"]
    long_su = {9, 11, 12}
    for r in range(9, 13):
        copy_row_style(su, 3 if r in long_su else 7, r, 7)

    copy_row_style(wb["Retrospection"], 3, 4, 9)


def fill_agile(path):
    wb = openpyxl.load_workbook(path)
    pb = wb["Product Backlog"]
    start = 22
    for i, row_data in enumerate(PRODUCT_BACKLOG_M3):
        write_row(pb, start + i, row_data)

    sb = wb["Sprint Backlog"]
    sprint3_header = 31
    write_row(sb, sprint3_header, ["SPRINT  3  BACKLOG"])
    task_start = sprint3_header + 1
    for i, task in enumerate(SPRINT3_TASKS):
        us, tid, desc, start_d, end_d, activity, status, hours = task[:8]
        d1 = task[8] if len(task) > 8 else 0
        d2 = task[9] if len(task) > 9 else 0
        row = task_start + i
        values = [us, tid, desc, start_d, end_d, ASSIGNEE, activity, status, hours, d1, d2]
        write_row(sb, row, values)
        # fill remaining day columns with 0 if needed
        for col in range(11, 20):
            if sb.cell(row, col).value is None:
                sb.cell(row, col, 0)

    # Update effort summary row (row 3): add sprint 3 hours
    m3_hours = sum(t[7] for t in SPRINT3_TASKS)
    old_total = sb.cell(3, 9).value or 98
    sb.cell(3, 9, old_total + m3_hours)

    standup = wb["Stand up Meeting"]
    su_start = 9  # first empty row after Sprint 2 stand-up entries
    for row_data in STANDUP_M3:
        write_row(standup, su_start, row_data)
        su_start += 1

    retro = wb["Retrospection"]
    r = RETRO_M3
    write_row(retro, 4, list(r))

    apply_m3_formatting(wb)
    wb.save(path)
    print(f"Updated {path}")


def fill_defects(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Defects"]
    row = 8  # first empty row after defect Sl 6
    for d in DEFECTS_M3:
        write_row(ws, row, d)
        row += 1
    long_rows = {8, 9, 10}
    for r in range(8, 13):
        copy_row_style(ws, 7 if r in long_rows else 2, r, 11)
    wb.save(path)
    print(f"Updated {path}")


def fill_unit_tests(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["UT"]
    row = 23  # first empty row after test Sl 21
    for t in UNIT_TESTS_M3:
        write_row(ws, row, t)
        row += 1
    long_rows = {23, 24, 25, 27, 28, 29, 30, 31, 35, 36, 37, 40, 41}
    last = ws.max_row
    for r in range(23, last + 1):
        if ws.cell(r, 1).value is None:
            continue
        copy_row_style(ws, 22 if r in long_rows else 2, r, 6)
    wb.save(path)
    print(f"Updated {path}")


if __name__ == "__main__":
    fill_agile(f"{ROOT}/Agile_Template_v1_Filled .xlsx")
    fill_defects(f"{ROOT}/Defect_Tracker_v1_Filled .xlsx")
    fill_unit_tests(f"{ROOT}/Unit_Test_Plan_v1_Filled (2) (1).xlsx")
    print("Done — Milestone 3 data added to all 3 documents.")
