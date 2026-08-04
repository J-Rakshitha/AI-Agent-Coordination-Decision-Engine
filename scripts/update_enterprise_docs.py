"""Append Enterprise 5-Phase items to Milestone 3 docs; fix dates; rebuild zip."""
from copy import copy
from datetime import datetime
from pathlib import Path
import zipfile

import openpyxl

DOCS = Path(__file__).resolve().parents[1] / "docs" / "Milestone_Documents"
ASSIGNEE = "J Rakshitha"
DATE_FMT = "d-mmm-yyyy"
RETRO_S3_START = datetime(2026, 7, 27)
RETRO_S3_END = datetime(2026, 8, 4)

PRODUCT_BACKLOG_E = [
    ("Sprint 3", "Sprint 3", "US033", "As a system, I want a Repository Discovery Agent performing real AST scans so that code symbols and complexity are indexed for conflict analysis.", "Must Have", "US013", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US034", "As a system, I want a Semantic Analysis Agent combining AST diff with Hybrid AI so that logic-level merge risks are explained.", "Must Have", "US033", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US035", "As a system, I want a Resolution Synthesizer Agent scoring three merge strategies so that the best coordination option is selected.", "Must Have", "US007", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US036", "As a system, I want a Quality Agent producing A/B/C grade scorecards from cyclomatic complexity metrics.", "Must Have", "US033", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US037", "As a system, I want semantic RAG search over the knowledge base using Gemini embeddings so that past patterns are retrieved by meaning.", "Must Have", "US011", ASSIGNEE, "Completed"),
    ("Sprint 3", "Sprint 3", "US038", "As a developer, I want README, pytest config, and milestone documents updated for enterprise features so that submission is complete.", "Should Have", "US037", ASSIGNEE, "Completed"),
]

SPRINT_TASKS_E = [
    ("US033", "T042", "Implement Repository Discovery Agent + code_parser AST service + CodeSymbol model", datetime(2026, 7, 27), datetime(2026, 7, 28), "Development", "Completed", 5, 5, 0),
    ("US034", "T043", "Implement Semantic Analysis Agent with AST diff + Hybrid AI prompts", datetime(2026, 7, 28), datetime(2026, 7, 29), "Development", "Completed", 5, 5, 0),
    ("US035", "T044", "Implement Resolution Synthesizer Agent with three scored strategies", datetime(2026, 7, 29), datetime(2026, 7, 30), "Development", "Completed", 4, 4, 0),
    ("US036", "T045", "Implement Quality Agent + evaluate_code_quality tool with A/B/C grading", datetime(2026, 7, 30), datetime(2026, 7, 31), "Development", "Completed", 4, 4, 0),
    ("US037", "T046", "Implement Knowledge Search RAG + embedding service + semantic_knowledge_search tool", datetime(2026, 7, 31), datetime(2026, 8, 1), "Development", "Completed", 5, 5, 0),
    ("US033", "T047", "Wire enterprise pipeline: Discovery → Semantic → Quality → Code Review → Notify", datetime(2026, 8, 1), datetime(2026, 8, 2), "Integration", "Completed", 4, 4, 0),
    ("US037", "T048", "Add Alembic migration 002_enterprise_intelligence + extend ConflictEvent JSON fields", datetime(2026, 8, 2), datetime(2026, 8, 3), "Development", "Completed", 3, 3, 0),
    ("US037", "T049", "Update DevCollab UI + Knowledge Base semantic search panel for enterprise outputs", datetime(2026, 8, 3), datetime(2026, 8, 4), "Development", "Completed", 4, 4, 0),
    ("US038", "T050", "Write test_enterprise_phase5.py (6 tests) + root pytest.ini; full suite 55 passed", datetime(2026, 8, 4), datetime(2026, 8, 4), "Testing", "Completed", 4, 4, 0),
    ("US038", "T051", "Update README, milestone Excel docs, git commit and push to GitHub", datetime(2026, 8, 4), datetime(2026, 8, 4), "Documentation", "Completed", 2, 2, 0),
]

DEFECT_E = (
    12, "J Rakshitha", datetime(2026, 8, 4),
    "GET /api/dev-collab/conflicts returned HTTP 500 — OperationalError: no such column conflict_events.discovery_context after enterprise model update.",
    "Sprint 3", "J Rakshitha", "Logical",
    "Applied Alembic migration 002_enterprise_intelligence; added root pytest.ini for async tests from project root.",
    datetime(2026, 8, 4), "Closed",
    "Conflicts API returns discovery_context, semantic_analysis, quality_report; 55 tests pass.",
)

UNIT_TESTS_E = [
    (41, "Repository Discovery indexes AST symbols", "POST /api/dev-collab/repository/discovery.", "Backend running.", "Response context.symbols_indexed > 0 and scan_source=local_ast.", "Pass — real AST scan verified."),
    (42, "Simulate conflict runs full enterprise pipeline", "POST simulate-demo-conflict; GET conflicts.", "Migration 002 applied.", "Conflict has discovery_context, semantic_analysis, quality_report, code_review_notes.", "Pass — 5-phase pipeline on detection."),
    (43, "Resolution synthesizer returns three strategies", "Simulate conflict; POST suggest-resolution.", "Conflict exists with semantic/quality data.", "Response includes synthesizer.options length 3 and best_strategy.", "Pass — synthesizer integrated."),
    (44, "Knowledge base semantic search (RAG)", "GET /api/system/knowledge-base/search?q=conflict merge resolution.", "Knowledge entries exist.", "Response includes results array and query echo.", "Pass — embedding search works."),
    (45, "Enterprise tools registered in tool registry", "GET /api/tools/.", "tool_handlers registered.", "Names include semantic_conflict_analyze, evaluate_code_quality, semantic_knowledge_search.", "Pass — 8 tools total."),
    (46, "Full regression suite after enterprise phases", "Run 'python -m pytest -v' from project root.", "pytest.ini at root with asyncio_mode=auto.", "55 tests pass with no failures.", "Pass — full suite green."),
]

STANDUP_E = [
    ("Sprint 3", "Day 8", "Conflict list API crashed after adding enterprise columns — SQLite missing discovery_context.", "Ran 'alembic upgrade head' for migration 002_enterprise_intelligence."),
    ("Sprint 3", "Day 9", "pytest from project root showed 48 async failures — backend/pytest.ini not loaded.", "Added root pytest.ini with asyncio_mode=auto and testpaths=backend/tests."),
]


def copy_row_style(ws, src_row, dst_row, max_col):
    src_h = ws.row_dimensions[src_row].height
    if src_h:
        ws.row_dimensions[dst_row].height = src_h
    for col in range(1, max_col + 1):
        src, dst = ws.cell(src_row, col), ws.cell(dst_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.alignment = copy(src.alignment)


def write_row(ws, row, values, start_col=1):
    for i, val in enumerate(values, start=start_col):
        ws.cell(row=row, column=i, value=val)


def set_date(cell, dt, ref):
    cell.value = dt
    cell.number_format = ref.number_format or DATE_FMT


def update_agile(path: Path):
    wb = openpyxl.load_workbook(path)
    pb = wb["Product Backlog"]
    pb_start = 34
    for i, row_data in enumerate(PRODUCT_BACKLOG_E):
        write_row(pb, pb_start + i, row_data)
        copy_row_style(pb, 22, pb_start + i, 8)

    sb = wb["Sprint Backlog"]
    task_start = 49
    for i, task in enumerate(SPRINT_TASKS_E):
        us, tid, desc, start_d, end_d, activity, status, hours = task[:8]
        d1 = task[8] if len(task) > 8 else 0
        d2 = task[9] if len(task) > 9 else 0
        row = task_start + i
        write_row(sb, row, [us, tid, desc, start_d, end_d, ASSIGNEE, activity, status, hours, d1, d2])
        copy_row_style(sb, 32, row, 23)
        for col in range(11, 20):
            if sb.cell(row, col).value is None:
                sb.cell(row, col, 0)

    extra_hours = sum(t[7] for t in SPRINT_TASKS_E)
    sb.cell(3, 9, (sb.cell(3, 9).value or 0) + extra_hours)

    # Fix all Sprint 3 task dates Jul 27 – Aug 4
    all_tasks = [
        ("T025", datetime(2026, 7, 27), datetime(2026, 7, 28)),
        ("T026", datetime(2026, 7, 28), datetime(2026, 7, 29)),
        ("T027", datetime(2026, 7, 29), datetime(2026, 7, 30)),
        ("T028", datetime(2026, 7, 30), datetime(2026, 7, 31)),
        ("T029", datetime(2026, 7, 31), datetime(2026, 8, 1)),
        ("T030", datetime(2026, 8, 1), datetime(2026, 8, 2)),
        ("T031", datetime(2026, 8, 2), datetime(2026, 8, 2)),
        ("T032", datetime(2026, 8, 2), datetime(2026, 8, 3)),
        ("T033", datetime(2026, 8, 3), datetime(2026, 8, 3)),
        ("T034", datetime(2026, 8, 3), datetime(2026, 8, 4)),
        ("T035", datetime(2026, 8, 4), datetime(2026, 8, 4)),
        ("T036", datetime(2026, 8, 4), datetime(2026, 8, 4)),
        ("T037", datetime(2026, 8, 4), datetime(2026, 8, 4)),
        ("T038", datetime(2026, 8, 4), datetime(2026, 8, 4)),
        ("T039", datetime(2026, 8, 4), datetime(2026, 8, 4)),
        ("T040", datetime(2026, 8, 4), datetime(2026, 8, 4)),
        ("T041", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ] + [(t[1], t[3], t[4]) for t in SPRINT_TASKS_E]
    task_map = {tid: (s, e) for tid, s, e in all_tasks}
    date_ref = sb.cell(21, 4)
    for row in range(32, task_start + len(SPRINT_TASKS_E)):
        tid = sb.cell(row, 2).value
        if tid in task_map:
            s, e = task_map[tid]
            set_date(sb.cell(row, 4), s, date_ref)
            set_date(sb.cell(row, 5), e, date_ref)

    standup = wb["Stand up Meeting"]
    su_row = 13
    for row_data in STANDUP_E:
        write_row(standup, su_row, row_data)
        copy_row_style(standup, 9, su_row, 7)
        su_row += 1

    retro = wb["Retrospection"]
    set_date(retro.cell(4, 3), RETRO_S3_START, retro.cell(2, 3))
    set_date(retro.cell(4, 4), RETRO_S3_END, retro.cell(2, 3))
    retro.cell(4, 9, "Delivered enterprise 5-phase intelligence (Discovery, Semantic, Synthesizer, Quality, RAG) with 55 automated tests and live multi-channel demo.")

    wb.save(path)
    print("Updated Agile:", path.name)


def update_defects(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Defects"]
    row = 13
    d = DEFECT_E
    write_row(ws, row, d)
    copy_row_style(ws, 8, row, 11)
    for r in range(2, ws.max_row + 1):
        for col in (3, 9):
            cell = ws.cell(r, col)
            if isinstance(cell.value, datetime):
                cell.number_format = DATE_FMT
    wb.save(path)
    print("Updated Defects:", path.name)


def update_unit_tests(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb["UT"]
    # Update regression test row 40 if present
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == 40:
            ws.cell(row, 3, "Run 'python -m pytest -v' from project root (or backend).")
            ws.cell(row, 5, "55 tests pass with no failures.")
            ws.cell(row, 6, "Pass — full suite green after enterprise phases.")
        if ws.cell(row, 2).value and "46+" in str(ws.cell(row, 2).value):
            ws.cell(row, 2, str(ws.cell(row, 2).value).replace("46+", "55"))

    row = 42
    for t in UNIT_TESTS_E:
        write_row(ws, row, t)
        copy_row_style(ws, 23, row, 6)
        row += 1
    wb.save(path)
    print("Updated Unit Tests:", path.name)


def rebuild_zip():
    finals = [
        "Agile_Template_v1_Filled.xlsx",
        "Defect_Tracker_v1_Filled.xlsx",
        "Unit_Test_Plan_v1_Filled.xlsx",
    ]
    zip_path = DOCS / "Milestone_Documents_M1_M2_M3.zip"
    zip_path.unlink(missing_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in finals:
            zf.write(DOCS / name, name)
    print("Rebuilt zip:", zip_path.name)


def cleanup_extras():
    keep = {
        "Agile_Template_v1_Filled.xlsx",
        "Defect_Tracker_v1_Filled.xlsx",
        "Unit_Test_Plan_v1_Filled.xlsx",
        "Milestone_Documents_M1_M2_M3.zip",
    }
    for p in DOCS.iterdir():
        if p.name in keep or p.name.startswith("~"):
            continue
        try:
            p.unlink()
            print("Removed:", p.name)
        except Exception as exc:
            print("Skip remove", p.name, exc)


if __name__ == "__main__":
    agile = DOCS / "Agile_Template_v1_Filled.xlsx"
    defects = DOCS / "Defect_Tracker_v1_Filled.xlsx"
    ut = DOCS / "Unit_Test_Plan_v1_Filled.xlsx"
    update_agile(agile)
    update_defects(defects)
    update_unit_tests(ut)
    cleanup_extras()
    rebuild_zip()
    print("Done — enterprise milestone docs updated.")
