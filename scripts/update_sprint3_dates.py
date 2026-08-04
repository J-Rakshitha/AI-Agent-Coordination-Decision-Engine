"""Update Sprint 3 dates to 24-Jul-2026 through 04-Aug-2026 across milestone documents."""
from datetime import datetime
import zipfile
import openpyxl
from pathlib import Path

DOCS = Path(__file__).resolve().parents[1] / "docs" / "Milestone_Documents"
SPRINT_START = datetime(2026, 7, 24)
SPRINT_END = datetime(2026, 8, 4)

# T025–T041: (task_id, start, end)
TASK_DATES = [
    ("T025", datetime(2026, 7, 24), datetime(2026, 7, 25)),
    ("T026", datetime(2026, 7, 25), datetime(2026, 7, 26)),
    ("T027", datetime(2026, 7, 26), datetime(2026, 7, 27)),
    ("T028", datetime(2026, 7, 27), datetime(2026, 7, 28)),
    ("T029", datetime(2026, 7, 28), datetime(2026, 7, 29)),
    ("T030", datetime(2026, 7, 29), datetime(2026, 7, 30)),
    ("T031", datetime(2026, 7, 30), datetime(2026, 7, 30)),
    ("T032", datetime(2026, 7, 31), datetime(2026, 8, 1)),
    ("T033", datetime(2026, 8, 1), datetime(2026, 8, 2)),
    ("T034", datetime(2026, 8, 2), datetime(2026, 8, 3)),
    ("T035", datetime(2026, 8, 3), datetime(2026, 8, 3)),
    ("T036", datetime(2026, 8, 3), datetime(2026, 8, 4)),
    ("T037", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T038", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T039", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T040", datetime(2026, 8, 4), datetime(2026, 8, 4)),
    ("T041", datetime(2026, 8, 4), datetime(2026, 8, 4)),
]

# Defect Sl: (submitted, action_taken)
DEFECT_DATES = {
    7: (datetime(2026, 8, 1), datetime(2026, 8, 1)),
    8: (datetime(2026, 7, 28), datetime(2026, 7, 28)),
    9: (datetime(2026, 8, 3), datetime(2026, 8, 3)),
    10: (datetime(2026, 7, 30), datetime(2026, 7, 30)),
    11: (datetime(2026, 8, 4), datetime(2026, 8, 4)),
}


def save_wb(wb, path):
    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    try:
        if path.exists():
            path.unlink()
        tmp.rename(path)
        print(f"Updated {path.name}")
        return True
    except PermissionError:
        tmp.rename(path.with_name(path.stem + "_dated.xlsx"))
        print(f"LOCKED — saved {path.stem}_dated.xlsx (close Excel and re-run)")
        return False


def update_agile(path):
    wb = openpyxl.load_workbook(path)
    retro = wb["Retrospection"]
    retro.cell(4, 3, SPRINT_START)
    retro.cell(4, 4, SPRINT_END)

    sb = wb["Sprint Backlog"]
    task_map = {tid: (s, e) for tid, s, e in TASK_DATES}
    for row in range(32, 49):
        tid = sb.cell(row, 2).value
        if tid in task_map:
            start, end = task_map[tid]
            sb.cell(row, 4, start)
            sb.cell(row, 5, end)

    return save_wb(wb, path)


def update_defects(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Defects"]
    for row in range(2, ws.max_row + 1):
        sl = ws.cell(row, 1).value
        if sl in DEFECT_DATES:
            submitted, action = DEFECT_DATES[sl]
            ws.cell(row, 3, submitted)
            ws.cell(row, 9, action)
    return save_wb(wb, path)


def rebuild_zip():
    zip_path = DOCS / "Milestone_Documents_M1_M2_M3.zip"
    zip_path.unlink(missing_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for name in [
            "Agile_Template_v1_Filled.xlsx",
            "Defect_Tracker_v1_Filled.xlsx",
            "Unit_Test_Plan_v1_Filled.xlsx",
        ]:
            dated = DOCS / name.replace(".xlsx", "_dated.xlsx")
            src = dated if dated.exists() else DOCS / name
            zf.write(src, name)
    print("Rebuilt zip")


if __name__ == "__main__":
    ag = update_agile(DOCS / "Agile_Template_v1_Filled.xlsx")
    df = update_defects(DOCS / "Defect_Tracker_v1_Filled.xlsx")
    rebuild_zip()
    print(f"Sprint 3 dates set to {SPRINT_START.date()} – {SPRINT_END.date()}")
